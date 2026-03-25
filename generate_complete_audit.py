import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from data_fetcher import fetch_dashboard_data
from logic_engine import calculate_kpis

# --- CONFIG ---
TEMPLATE_PATH = r"c:\Users\ELCOT\Downloads\injini-mel-dashboard-main\injini-mel-dashboard-main\Full_Dashboard_All_Indicators.xlsx"
OUTPUT_PATH = r"c:\Users\ELCOT\Downloads\injini-mel-dashboard-main\injini-mel-dashboard-main\Dashboard_Audit_Report.xlsx"

def format_signed(val):
    try:
        f = float(val)
        return f"+{f:,.0f}" if f >= 0 else f"{f:,.0f}"
    except:
        return str(val)

def generate_audit():
    print("Fetching data from Airtable...")
    # Fetch data
    raw_df = fetch_dashboard_data()
    print("Calculating KPIs...")
    kpis = calculate_kpis(raw_df)
    
    print(f"Loading template: {TEMPLATE_PATH}")
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    def find_row_by_label(ws, label, col=1, start_r=1):
        for r in range(start_r, 1000):
            val = ws.cell(row=r, column=col).value
            if val and label.lower() in str(val).lower(): return r
        return -1

    # 1. 📊 Master Summary
    if '📊 Master Summary' in wb.sheetnames:
        ws = wb['📊 Master Summary']
        mapping = {'Cohort 1': 4, 'Cohort 2': 5, 'Cohort 3': 6, 'Programme': 7}
        sales_row = find_row_by_label(ws, "Sales Growth", 1, 1)
        c_sales_row = find_row_by_label(ws, "Cohort", 2, sales_row)
        p_row = find_row_by_label(ws, "Profit Growth", 1, 1)
        c_p_row = find_row_by_label(ws, "Cohort", 2, p_row)
        for ch, col in mapping.items():
            if ch == 'Programme':
                ws.cell(row=c_sales_row, column=col).value = f"{kpis['Program_Overview'].get('Average Sales Growth', 'N/A')}%"
                ws.cell(row=c_p_row, column=col).value = f"{kpis['Program_Overview'].get('Average Profit Growth', 'N/A')}%"
                prog_sales_row = find_row_by_label(ws, "Programme", 2, sales_row)
                prog_p_row = find_row_by_label(ws, "Programme", 2, p_row)
                if prog_sales_row != -1: ws.cell(row=prog_sales_row, column=col).value = f"{kpis['Program_Overview'].get('Average Sales Growth', 'N/A')}%"
                if prog_p_row != -1: ws.cell(row=prog_p_row, column=col).value = f"{kpis['Program_Overview'].get('Average Profit Growth', 'N/A')}%"
            else:
                det = kpis['Cohort_Detail'].get(ch, {})
                if c_sales_row != -1: ws.cell(row=c_sales_row, column=col).value = f"{det.get('cohort_median_sg', 'N/A')}%"
                if c_p_row != -1: ws.cell(row=c_p_row, column=col).value = f"{det.get('cohort_median_pg', 'N/A')}%"

    def get_cohort_sections(ws):
        sections = {}
        for r in range(1, 1000):
            val = ws.cell(row=r, column=1).value
            if val and "cohort" in str(val).lower() and "|" in str(val):
                p = str(val).split('|')[0].lower()
                if "cohort 1" in p: sections["Cohort 1"] = r
                elif "cohort 2" in p: sections["Cohort 2"] = r
                elif "cohort 3" in p: sections["Cohort 3"] = r
                elif "cohort 4" in p: sections["Cohort 4"] = r
        return sections

    def fill_v_sheet(sheet_name, cohort_key, data_list, col_count):
        if sheet_name not in wb.sheetnames: return
        ws = wb[sheet_name]
        sections = get_cohort_sections(ws)
        if cohort_key not in sections: return
        start_r = sections[cohort_key] + 2
        
        # AGGRESSIVE CLEAR
        curr = start_r
        while curr < start_r + 50:
            v_a = ws.cell(row=curr, column=1).value
            is_m = any(isinstance(ws.cell(row=curr, column=c), openpyxl.cell.cell.MergedCell) for c in range(1, col_count+1))
            label = str(v_a).lower() if v_a else ""
            if is_m or ("total" in label and "cohort" in label): break
            for c in range(1, col_count + 1): 
                ws.cell(row=curr, column=c).value = None
            curr += 1
            
        # Write
        for i, rowdata in enumerate(data_list):
            r = start_r + i
            if any(isinstance(ws.cell(row=r, column=c), openpyxl.cell.cell.MergedCell) for c in range(1, col_count+1)):
                ws.insert_rows(r)
            for j, val in enumerate(rowdata):
                ws.cell(row=r, column=j+1).value = val

    # Fill sheets
    for ch in ['Cohort 1', 'Cohort 2', 'Cohort 3', 'Cohort 4']:
        cd = kpis['Cohort_Detail'].get(ch, {})
        fill_v_sheet('📈 Sales Growth', ch, [[g['name'], g['months'], g['rule'], g['base_val'], g['recent_val'], f"(R{g['recent_val']:,.0f}-R{g['base_val']:,.0f})/R{g['base_val']:,.0f}", f"{g['sales_growth']}%" if isinstance(g['sales_growth'],(int,float)) else g['sales_growth'], "Positive" if (isinstance(g['sales_growth'],(int,float)) and g['sales_growth']>=0) else "Negative" if isinstance(g['sales_growth'],(int,float)) else ""] for g in cd.get('growth_table',[])], 8)
        fill_v_sheet('💰 Profit Growth', ch, [[g['name'], g['months'], g['rule'], g['profit_base_val'], g['profit_recent_val'], f"(R{g['profit_recent_val']:,.0f}-R{g['profit_base_val']:,.0f})/ABS(R{g['profit_base_val']:,.0f})", f"{g['profit_growth']}%" if isinstance(g['profit_growth'],(int,float)) else g['profit_growth'], "Positive" if (isinstance(g['profit_growth'],(int,float)) and g['profit_growth']>=0) else "Negative" if isinstance(g['profit_growth'],(int,float)) else ""] for g in cd.get('growth_table',[])], 8)
        fill_v_sheet('👷 Jobs Indicators', ch, [[j['name'], j['baseline'], j['total'], j['total'], format_signed(j['new']), format_signed(j['new_female']), format_signed(j['new_youth']), f"{j['pct_change']}%", j['baseline_female'], j['baseline_youth']] for j in cd.get('jobs_table',[])], 10)
        fill_v_sheet('🌍 Reach Indicators', ch, [[u['name'], u['tot_learners'], u['tot_educators'], u['tot_learners']+u['tot_educators'], u['new_learners'], u['new_educators'], u['new_learners']+u['new_educators'], u['new_learners'], u['schools'], f"SA:{u['sa_schools']}/Q1-3:{u['q13_schools']}"] for u in cd.get('users_table',[])], 10)
        fill_v_sheet('👩‍🎓 Disaggregation', ch, [[d['name'], f"{d['female_stu']}%", f"{d['female_tea']}%", f"{d['rural_stu']}%", f"{d['rural_tea']}%", f"{d['disability_stu']}%", f"{d['disability_tea']}%"] for d in cd.get('disaggregation',[])], 7)

    # Investments
    if '💳 Investments' in wb.sheetnames:
        ws = wb['💳 Investments']
        s_row = find_row_by_label(ws, "Cohort", 1, 1) + 1
        if s_row == 0: s_row = 4
        
        c_clear = s_row
        while c_clear < s_row + 50:
            is_m = any(isinstance(ws.cell(row=c_clear, column=c), openpyxl.cell.cell.MergedCell) for c in range(1, 6))
            v_a = ws.cell(row=c_clear, column=1).value
            if is_m or (v_a and "total" in str(v_a).lower()): break
            for c in range(1, 6): ws.cell(row=c_clear, column=c).value = None
            c_clear += 1
        
        row_i = s_row
        for ch in ['Cohort 1', 'Cohort 2', 'Cohort 3', 'Cohort 4']:
            invs = kpis['Cohort_Detail'].get(ch, {}).get('investments_table', [])
            items = invs if invs else [{"name":"", "month":"", "value":0, "investor":"No grants reported"}]
            for inv in items:
                if any(isinstance(ws.cell(row=row_i, column=cc), openpyxl.cell.cell.MergedCell) for cc in range(1, 6)): ws.insert_rows(row_i)
                ws.cell(row=row_i, column=1).value = ch if (not invs or inv==items[0]) else ""
                if invs:
                    ws.cell(row=row_i, column=2).value = inv['name']
                    ws.cell(row=row_i, column=3).value = inv['month']
                    ws.cell(row=row_i, column=4).value = f"R {inv['value']:,.0f}"
                else: ws.cell(row=row_i, column=4).value = "R 0"
                ws.cell(row=row_i, column=5).value = inv['investor']
                row_i += 1

    wb.save(OUTPUT_PATH)
    print(f"Audit Report saved to: {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_audit()
